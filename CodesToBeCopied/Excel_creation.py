import pandas as pd
from errno import EACCES, EPERM, ENOENT
from operator import itemgetter
from openpyxl import load_workbook


def print_error_message(e, file_name):
    #PermissionError
	if e.errno==EPERM or e.errno==EACCES:
		print("Error")
		print("PermissionError error({0}): {1} for:\n{2}".format(e.errno, e.strerror, file_name))
		#FileNotFoundError
	elif e.errno==ENOENT:
		print("FileNotFoundError error({0}): {1} as:\n{2}".format(e.errno, e.strerror, file_name))
	elif IOError:
		print("I/O error({0}): {1} as:\n{2}".format(e.errno, e.strerror, file_name))
	elif OSError:
		print("OS error({0}): {1} as:\n{2}".format(e.errno, e.strerror, file_name))
		
def scan_application_objects_list(application_object_list,b):
	Folders = {}
	Connections = []
	objects = []
	
	with open(application_object_list,"r") as f:
		#lets read it into an array
		lines = f.read().splitlines()
		
		for i,line in enumerate(lines):
			#print(line)
			object = line.split("\t")
			temp = object[0].split('/')
			temp.append(object[1])
			if(temp[0]=="ConnectInfoProject"):
				Connections.append(temp[1])
			else:
				if(temp[0] not in Folders):
					Folders[temp[0]] = 0
				objects.append(temp)
		
	#print(Folders)
	#print(Connections)
	#print(objects)
	
	objects = sorted(objects, key=itemgetter(2), reverse=False)
	
	for i,n in enumerate(objects):
		if n[2] == "RelationalDataObject":
			n[2] = "Relational Data Object"
		elif n[2] == "FlatFileDataObject":
			n[2] = "Flat File Data Object"

	#print(objects)
	col_1=[]
	col_2=[]
	col_3=[]
	col_4 = []
	col_5 = []
	for i in objects:
		col_1.append(i[0])
		col_2.append(i[1])
		col_3.append(i[2])
		col_4.append("reuse")
		col_5.append(i[1])
	
	for i in Connections:
		col_1.append("-")
		col_2.append(i)
		col_3.append("Connection")
		col_4.append("reuse")
		col_5.append(i)
	
	
	wb = load_workbook(b)
	mysheet = wb.create_sheet("Resolution Specification")

	mysheet.append(["Folder Name","Source Object Name","Object Type","Object Resolution","Target Object Name"])
	for i in range(len(col_1)):
		#print([col_1[i],col_2[i],col_3[i],col_4[i],col_5[i]])
		mysheet.append([col_1[i],col_2[i],col_3[i],col_4[i],col_5[i]])
	
	for col in mysheet.columns:
		max_length = 0
		column = col[0].column # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		mysheet.column_dimensions[column].width = adjusted_width
	
	#dv = DataValidation(type="list", formula1='"reuse,replace,rename,none"', allow_blank=True,showDropDown = True)
	# Optionally set a custom error message
	#dv.error ='Your entry is not in the list'
	#dv.errorTitle = 'Invalid Entry'
	
	# Optionally set a custom prompt message
	#dv.prompt = 'Please select from the list'
	#dv.promptTitle = 'List Selection'
	
	#mysheet.add_data_validation(dv)
	wb.save(b)
	
